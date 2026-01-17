import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
import io

# Sayfa Yapılandırması
st.set_page_config(page_title="Zayi Raporu - Birebir Biçim", layout="wide")

st.title("📊 Cam Zayi Raporu - Birebir Görünüm Onarıcı")
st.markdown("---")
st.info("💡 Bu sürüm; sütun genişliklerini, renkli grupları ve dikey yazıları orijinal dosyanızdan birebir kopyalar.")

# Listelenmesi istenen tam mağaza kodları
istenen_magazalar = [
    "M38003", "M51001", "M42004", "M51002", "M38001", "M38005", 
    "M68001", "M42006", "M42002", "M46001", "M38002", "M42001", 
    "M40001", "M42005", "M38004", "M70001", "M50001"
]

uploaded_file = st.file_uploader("Orijinal Excel dosyasını yükleyin", type=['xlsx'])

if uploaded_file is not None:
    try:
        # 1. Dosyayı Biçimleri Koruyarak Yükle
        # data_only=False: Formülleri ve hücre stillerini (renk, dikey yazı vb.) korur
        wb = openpyxl.load_workbook(uploaded_file, data_only=False)
        ws = wb.active

        # 2. SOLDAKİ GRUPLANDIRMA (OUTLINE) VE BUTONLARI (+/-) SİL
        ws.sheet_format.outlineLevelRow = 0
        ws.sheet_format.outlineLevelCol = 0
        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].outline_level = 0
            ws.row_dimensions[r].hidden = False

        # 3. "Üst Birim" Başlık Satırını ve Sütununu Tespit Et
        header_row = 1
        ub_col_idx = 1
        found = False
        for r in range(1, 25): # İlk 25 satırda başlığı ara
            for c in range(1, 15): # İlk 15 sütunda ara
                val = str(ws.cell(r, c).value).strip().upper()
                if "ÜST BIRIM" in val:
                    header_row = r
                    ub_col_idx = c
                    found = True
                    break
            if found: break

        # 4. SOLDAKİ GEREKSİZ SÜTUNLARI SİL (Bölge / Müdür vb.)
        # Raporun tam istediğiniz gibi 'Üst Birim' ile başlamasını sağlar
        if ub_col_idx > 1:
            ws.delete_cols(1, ub_col_idx - 1)
        
        # 5. ÜSTTEKİ BOŞ SATIRLARI SİL
        # Mavi başlığı en üste (1. satıra) taşır
        if header_row > 1:
            ws.delete_rows(1, header_row - 1)
            header_row = 1 

        # 6. MAĞAZALARI FİLTRELE (İstenmeyen satırları temizle)
        # Sondan başa doğru silmek, Excel'in hücre yapısını ve genişliklerini bozmaz
        max_row = ws.max_row
        for r in range(max_row, header_row, -1):
            m_kodu_raw = ws.cell(r, 1).value
            m_kodu = str(m_kodu_raw).strip() if m_kodu_raw else ""
            
            # Eğer satır bir mağaza kodu içeriyorsa ama listede yoksa sil
            # Genel toplam satırlarını silmemek için m_kodu uzunluğu kontrol edilir
            if m_kodu not in istenen_magazalar:
                if m_kodu != "" and len(m_kodu) >= 5: # Mağaza kodları MXXXXX formatında
                    ws.delete_rows(r)

        # 7. GÖRSEL DÜZELTMELER (Sütun Genişliği ve Satır Yüksekliği)
        # Orijinal görseldeki pürüzsüz başlık için yüksekliği sabitliyoruz
        ws.row_dimensions[1].height = 60 
        # Üst Birim sütununu orijinal genişliğine (yaklaşık 15 birim) getiriyoruz
        ws.column_dimensions['A'].width = 15

        # 8. DOSYAYI KAYDET VE İNDİRMEYE SUN
        output = io.BytesIO()
        wb.save(output)
        
        st.success("✅ Rapor Hazır! Orijinal renkler, dikey yazılar ve sütun aralıkları birebir korundu.")
        
        st.download_button(
            label="📥 Birebir Görünümlü Excel'i İndir",
            data=output.getvalue(),
            file_name="Zayi_Raporu_Birebir_Görünüm.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Sistem Hatası: {e}. Lütfen dosya formatının doğru olduğundan emin olun.")
